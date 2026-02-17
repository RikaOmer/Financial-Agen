import openpyxl
import os
import re
from datetime import datetime
from collections import defaultdict

data_dir = "Data"
all_transactions = []

def parse_amount(val):
    """Parse amount from various formats like '45.00 ₪', '16.25 $', or numeric"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace('\u20aa', '').replace('$', '').replace(',', '').replace('\u200f', '').strip()
    neg = False
    if s.startswith('-'):
        neg = True
        s = s[1:].strip()
    try:
        v = float(s)
        return -v if neg else v
    except:
        return 0

def parse_date_dmy(val):
    """Parse date from DD/MM/YYYY or DD-MM-YYYY or datetime"""
    if isinstance(val, datetime):
        return val
    if val is None:
        return None
    s = str(val).strip()
    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
        try:
            return datetime.strptime(s, fmt)
        except:
            continue
    return None

# ============================================================
# 1. Parse Visa credit card statements (פירוט עסקאות files)
# ============================================================
visa_files = [f for f in os.listdir(data_dir) if f.startswith('\u05e4\u05d9\u05e8\u05d5\u05d8 \u05e2\u05e1\u05e7\u05d0\u05d5\u05ea')]
for fname in sorted(visa_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        card_info = str(ws.cell(2, 1).value or '')
        card_num = '4581' if '4581' in card_info else ('3147' if '3147' in card_info else 'unknown')

        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_str = str(row[0]) if row and row[0] else ''
            if '\u05ea\u05d0\u05e8\u05d9\u05da' in row_str:
                header_row = i
                break

        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] is None or str(row[0]).strip() == '':
                continue
            row0_str = str(row[0]).strip()
            if '\u05e1\u05d4"\u05db' in row0_str or '\u05e1\u05db\u05d5\u05dd' in row0_str:
                continue

            date = parse_date_dmy(row[0])
            name = str(row[1] or '').strip()
            orig_amount = parse_amount(row[2])
            charge_amount = parse_amount(row[3])
            installment_info = str(row[4] or '').strip() if len(row) > 4 else ''

            if date and name and charge_amount > 0:
                all_transactions.append({
                    'date': date,
                    'name': name,
                    'amount': charge_amount,
                    'orig_amount': orig_amount,
                    'source': f'Visa-{card_num}',
                    'installment': installment_info,
                    'file': fname,
                    'sheet': sheet_name,
                    'category': ''
                })

# ============================================================
# 2. Parse Mastercard files (כרטיס מאסטרקארד 4045)
# ============================================================
mc_files = [f for f in os.listdir(data_dir) if '\u05de\u05d0\u05e1\u05d8\u05e8\u05e7\u05d0\u05e8\u05d3' in f]
for fname in sorted(mc_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_str = str(row[0]) if row and row[0] else ''
            if '\u05ea\u05d0\u05e8\u05d9\u05da' in row_str:
                header_row = i
                break

        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] is None:
                continue

            date = parse_date_dmy(row[0])
            name = str(row[1] or '').strip()
            orig_amount = parse_amount(row[2])
            charge_amount = parse_amount(row[3])
            category = str(row[5] or '').strip() if len(row) > 5 else ''

            if not charge_amount:
                charge_amount = orig_amount

            if date and name and charge_amount > 0:
                all_transactions.append({
                    'date': date,
                    'name': name,
                    'amount': charge_amount,
                    'orig_amount': orig_amount,
                    'source': 'Mastercard-4045',
                    'installment': '',
                    'category': category,
                    'file': fname,
                    'sheet': sheet_name
                })

# ============================================================
# 3. Parse bank debit transactions (transaction-details_export)
# ============================================================
bank_files = [f for f in os.listdir(data_dir) if f.startswith('transaction-details')]
for fname in sorted(bank_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            row0_str = str(row[0]).strip()
            if '\u05e1\u05d4"\u05db' in row0_str or row0_str == '':
                continue

            date = parse_date_dmy(row[0])
            desc = str(row[2] or '').strip()
            amount = parse_amount(row[5])
            installment_info = str(row[10] or '').strip() if len(row) > 10 else ''

            if date and desc and amount > 0:
                all_transactions.append({
                    'date': date,
                    'name': desc,
                    'amount': amount,
                    'orig_amount': amount,
                    'source': 'Bank-Debit',
                    'installment': installment_info,
                    'file': fname,
                    'sheet': sheet_name,
                    'category': ''
                })

# ============================================================
# Now analyze
# ============================================================

print(f"Total raw transactions loaded: {len(all_transactions)}")
source_counts = defaultdict(int)
for t in all_transactions:
    source_counts[t['source']] += 1
print(f"Sources: {dict(source_counts)}")

# Group by month
by_month = defaultdict(list)
for t in all_transactions:
    key = t['date'].strftime('%Y-%m')
    by_month[key].append(t)

print(f"\nMonths found: {sorted(by_month.keys())}")
for m in sorted(by_month.keys()):
    txns = by_month[m]
    total = sum(t['amount'] for t in txns)
    print(f"  {m}: {len(txns)} transactions, total {total:,.2f} NIS")

# Print all transactions sorted by month and amount
print("\n" + "=" * 100)
print("ALL TRANSACTIONS BY MONTH")
print("=" * 100)
for m in sorted(by_month.keys()):
    txns = sorted(by_month[m], key=lambda x: -x['amount'])
    total = sum(t['amount'] for t in txns)
    print(f"\n{'=' * 80}")
    print(f"  MONTH: {m}  |  Transactions: {len(txns)}  |  Total: {total:,.2f} NIS")
    print(f"{'=' * 80}")
    for t in txns:
        inst = f" [{t['installment']}]" if t['installment'] else ""
        cat = f" ({t['category']})" if t.get('category') else ""
        print(f"  {t['date'].strftime('%d/%m/%Y')}  {t['amount']:>10,.2f} NIS  {t['name'][:50]:<50s} [{t['source']}]{inst}{cat}")
