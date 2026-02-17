import openpyxl
import os
import re
import sys
from datetime import datetime
from collections import defaultdict

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

data_dir = "Data"
all_transactions = []

def parse_amount(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = s.replace('\u20aa', '').replace('$', '').replace(',', '').replace('\u200f', '').strip()
    neg = False
    if s.startswith('-'):
        neg = True
        s = s[1:].strip()
    try:
        v = float(s)
        return -v if neg else v
    except:
        return 0

def parse_date_dmy(val):
    if isinstance(val, datetime):
        return val
    if val is None:
        return None
    s = str(val).strip()
    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
        try:
            return datetime.strptime(s, fmt)
        except:
            continue
    return None

# ============================================================
# Parse all files
# ============================================================

# 1. Visa credit card (פירוט עסקאות)
visa_files = [f for f in os.listdir(data_dir) if f.startswith('\u05e4\u05d9\u05e8\u05d5\u05d8 \u05e2\u05e1\u05e7\u05d0\u05d5\u05ea')]
for fname in sorted(visa_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        card_info = str(ws.cell(2, 1).value or '')
        card_num = '4581' if '4581' in card_info else ('3147' if '3147' in card_info else 'unknown')

        # Get billing date from row 1
        billing_info = str(ws.cell(1, 1).value or '')

        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_str = str(row[0]) if row and row[0] else ''
            if '\u05ea\u05d0\u05e8\u05d9\u05da' in row_str:
                header_row = i
                break
        if header_row is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] is None or str(row[0]).strip() == '':
                continue
            row0_str = str(row[0]).strip()
            if '\u05e1\u05d4"\u05db' in row0_str or '\u05e1\u05db\u05d5\u05dd' in row0_str:
                continue
            date = parse_date_dmy(row[0])
            name = str(row[1] or '').strip()
            orig_amount = parse_amount(row[2])
            charge_amount = parse_amount(row[3])
            installment_info = str(row[4] or '').strip() if len(row) > 4 else ''
            if date and name and charge_amount > 0:
                all_transactions.append({
                    'date': date, 'name': name, 'amount': charge_amount,
                    'orig_amount': orig_amount, 'source': f'Visa-{card_num}',
                    'installment': installment_info, 'category': '',
                    'billing': billing_info
                })

# 2. Mastercard (מאסטרקארד 4045)
mc_files = [f for f in os.listdir(data_dir) if '\u05de\u05d0\u05e1\u05d8\u05e8\u05e7\u05d0\u05e8\u05d3' in f]
for fname in sorted(mc_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            row_str = str(row[0]) if row and row[0] else ''
            if '\u05ea\u05d0\u05e8\u05d9\u05da' in row_str:
                header_row = i
                break
        if header_row is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if row[0] is None:
                continue
            date = parse_date_dmy(row[0])
            name = str(row[1] or '').strip()
            orig_amount = parse_amount(row[2])
            charge_amount = parse_amount(row[3])
            category = str(row[5] or '').strip() if len(row) > 5 else ''
            if not charge_amount:
                charge_amount = orig_amount
            if date and name and charge_amount > 0:
                all_transactions.append({
                    'date': date, 'name': name, 'amount': charge_amount,
                    'orig_amount': orig_amount, 'source': 'Mastercard-4045',
                    'installment': '', 'category': category, 'billing': ''
                })

# 3. Bank debit (transaction-details_export)
bank_files = [f for f in os.listdir(data_dir) if f.startswith('transaction-details')]
for fname in sorted(bank_files):
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            row0_str = str(row[0]).strip()
            if '\u05e1\u05d4"\u05db' in row0_str or row0_str == '':
                continue
            date = parse_date_dmy(row[0])
            desc = str(row[2] or '').strip()
            amount = parse_amount(row[5])
            installment_info = str(row[10] or '').strip() if len(row) > 10 else ''
            if date and desc and amount > 0:
                all_transactions.append({
                    'date': date, 'name': desc, 'amount': amount,
                    'orig_amount': amount, 'source': 'Bank-Debit',
                    'installment': installment_info, 'category': '', 'billing': ''
                })

# ============================================================
# Classify each transaction
# ============================================================

def classify(t):
    """Classify transaction as: essential, avoidable, or reducible"""
    name = t['name'].lower()
    cat = t.get('category', '').lower()
    amount = t['amount']
    installment = t.get('installment', '')

    # FIXED COMMITMENTS / ESSENTIALS (can't easily avoid)
    essentials = [
        # Installment payments (loans, furniture etc.)
        ('\u05ea\u05e9\u05dc\u05d5\u05dd', 'installment'),
        ('\u05de\u05ea\u05d5\u05da', 'installment'),
        # Rent / Housing
        ('\u05e9\u05db\u05e8 \u05d3\u05d9\u05e8\u05d4', 'rent'),
        # Insurance
        ('\u05d1\u05d9\u05d8\u05d5\u05d7', 'insurance'),
        ('passportcard', 'travel_insurance'),
        # Utilities / phone
        ('019', 'phone_bill'),
        ('\u05e1\u05dc\u05d5\u05dc\u05e8', 'phone'),
        # Health
        ('\u05de\u05db\u05d1\u05d9', 'health'),
        ('\u05e7\u05d5\u05e4\u05ea \u05d7\u05d5\u05dc\u05d9\u05dd', 'health'),
        # Supermarket / groceries
        ('\u05e9\u05d5\u05e4\u05e8\u05e1\u05dc', 'groceries'),
        ('\u05e8\u05de\u05d9 \u05dc\u05d5\u05d9', 'groceries'),
        ('\u05d5\u05d9\u05e7\u05d8\u05d5\u05e8\u05d9', 'groceries'),
        ('\u05d9\u05d5\u05d7\u05e0\u05e0\u05d9', 'groceries'),
        ('\u05d0\u05dd \u05e4\u05d9 \u05d0\u05dd', 'groceries'),
        # Gym
        ('black iron', 'gym'),
        ('\u05d7\u05d3\u05e8 \u05db\u05d5\u05e9\u05e8', 'gym'),
        # Gas
        ('\u05d3\u05dc\u05e7', 'gas'),
        ('\u05e1\u05d5\u05e0\u05d5\u05dc', 'gas'),
        ('\u05e4\u05d6 \u05d2\u05d6', 'gas'),
        # Transport essentials
        ('\u05e8\u05db\u05d1\u05ea', 'transport'),
        # Pet
        ('\u05d7\u05ea\u05d5\u05dc', 'pet'),
        # Medical
        ('\u05e8\u05d5\u05e4\u05d0', 'medical'),
        ('\u05de\u05e8\u05e4\u05d0\u05d4', 'medical'),
    ]

    # SUBSCRIPTIONS (could cancel some)
    subscriptions = [
        ('spotify', 'music_sub', 'Could downgrade to free tier'),
        ('amazon prime', 'streaming_sub', 'Consider if you use it enough'),
        ('cursor', 'dev_tool_sub', 'Dev tool - evaluate if needed'),
        ('claude.ai', 'ai_sub', 'AI subscription'),
        ('anthropic', 'ai_sub', 'AI API usage'),
        ('google one', 'cloud_sub', 'Cloud storage - check if needed'),
        ('airalo', 'esim_sub', 'Travel eSIM'),
    ]

    # AVOIDABLE categories
    avoidable_keywords = [
        # Restaurants / eating out
        ('\u05de\u05e1\u05e2\u05d3', 'restaurant'),
        ('\u05e7\u05e4\u05d4', 'cafe'),
        ('cafe', 'cafe'),
        ('hong bao', 'restaurant'),
        ('\u05e4\u05d9\u05e6\u05e8\u05d9\u05d4', 'pizza'),
        ('\u05d1\u05d5\u05e8\u05d2\u05e8', 'fastfood'),
        ('sheldon', 'restaurant'),
        ('kisu', 'restaurant'),
        ('ore -', 'restaurant'),
        ('\u05d9\u05d4\u05d5\u05e9\u05e2', 'cafe'),
        ('\u05e7\u05e8\u05d9\u05d0\u05ea', 'cafe'),
        # Online shopping / impulse
        ('amazon mktpl', 'online_shopping'),
        ('amazon mktplace', 'online_shopping'),
        ('iherb', 'online_shopping'),
        ('next online', 'online_shopping'),
        ('hataco', 'shopping'),
        # Entertainment / leisure
        ('lime', 'scooter_rental'),
        # Alcohol / bars
        ('\u05d1\u05d9\u05e8\u05d4', 'alcohol'),
        ('\u05d0\u05dc\u05db\u05d5\u05d4\u05dc', 'alcohol'),
        # Takeout / delivery
        ('\u05d5\u05d5\u05dc\u05d8', 'delivery'),
        ('wolt', 'delivery'),
        # Impulse purchases
        ('duty free', 'duty_free'),
        ('king power', 'duty_free'),
        # Shopping
        ('\u05e7\u05e0\u05d9\u05d5\u05df', 'shopping'),
    ]

    # REDUCIBLE - things you need but could pay less for
    reducible_keywords = [
        ('\u05de\u05e1\u05e2\u05d3', 'eat_out_less'),
        ('\u05e7\u05e4\u05d4', 'coffee_less'),
    ]

    # Check installments first
    if installment and ('\u05ea\u05e9\u05dc\u05d5\u05dd' in installment or '\u05de\u05ea\u05d5\u05da' in installment or ' \u05de -' in installment):
        return 'essential', 'installment_payment', ''

    # Check subscriptions
    for kw, sub_cat, advice in subscriptions:
        if kw in name:
            return 'subscription', sub_cat, advice

    # Check avoidable
    for kw, av_cat in avoidable_keywords:
        if kw in name or kw in cat:
            return 'avoidable', av_cat, ''

    # Check category from Mastercard
    if cat:
        cat_lower = cat
        if '\u05de\u05e1\u05e2\u05d3' in cat_lower or '\u05d0\u05d5\u05db\u05dc' in cat_lower:
            return 'avoidable', 'restaurant', ''
        if '\u05e7\u05e0\u05d9\u05d5\u05ea' in cat_lower or '\u05e7\u05e0\u05d9\u05d5\u05df' in cat_lower:
            return 'avoidable', 'shopping', ''
        if '\u05d1\u05d9\u05d3\u05d5\u05e8' in cat_lower:
            return 'avoidable', 'entertainment', ''
        if '\u05d1\u05d9\u05d2\u05d5\u05d3' in cat_lower or '\u05d0\u05d5\u05e4\u05e0\u05d4' in cat_lower:
            return 'avoidable', 'fashion', ''

    return 'other', 'unclassified', ''


# ============================================================
# Analyze by month
# ============================================================

by_month = defaultdict(list)
for t in all_transactions:
    key = t['date'].strftime('%Y-%m')
    by_month[key].append(t)

# Focus on the main spending months
focus_months = ['2025-10', '2025-11', '2025-12', '2026-01', '2026-02']

output = []

def p(s=''):
    output.append(s)

p("=" * 90)
p("                    BANK SPENDING ANALYSIS - MONTHLY BREAKDOWN")
p("=" * 90)

grand_total = 0
grand_avoidable = 0
grand_subscription = 0

for m in focus_months:
    if m not in by_month:
        continue
    txns = by_month[m]
    total = sum(t['amount'] for t in txns)
    grand_total += total

    # Classify all transactions
    classified = []
    for t in txns:
        cls, sub_cls, advice = classify(t)
        classified.append({**t, 'cls': cls, 'sub_cls': sub_cls, 'advice': advice})

    essential_total = sum(t['amount'] for t in classified if t['cls'] == 'essential')
    avoidable_total = sum(t['amount'] for t in classified if t['cls'] == 'avoidable')
    subscription_total = sum(t['amount'] for t in classified if t['cls'] == 'subscription')
    other_total = sum(t['amount'] for t in classified if t['cls'] == 'other')

    grand_avoidable += avoidable_total
    grand_subscription += subscription_total

    month_names = {
        '2025-10': 'October 2025', '2025-11': 'November 2025',
        '2025-12': 'December 2025', '2026-01': 'January 2026',
        '2026-02': 'February 2026'
    }

    p(f"\n{'='*90}")
    p(f"  {month_names[m]}")
    p(f"{'='*90}")
    p(f"  TOTAL SPENT: {total:,.0f} NIS")
    p(f"  ├── Essentials/Installments: {essential_total:,.0f} NIS")
    p(f"  ├── Subscriptions:           {subscription_total:,.0f} NIS")
    p(f"  ├── Avoidable/Leisure:       {avoidable_total:,.0f} NIS")
    p(f"  └── Other/Unclassified:      {other_total:,.0f} NIS")

    # Show avoidable items
    avoidable_items = [t for t in classified if t['cls'] == 'avoidable']
    if avoidable_items:
        avoidable_items.sort(key=lambda x: -x['amount'])
        p(f"\n  THINGS YOU DIDN'T HAVE TO BUY / COULD PAY LESS:")
        p(f"  {'─'*80}")
        for t in avoidable_items:
            p(f"    {t['date'].strftime('%d/%m')}  {t['amount']:>8,.0f} NIS  {t['name'][:55]:<55s} [{t['sub_cls']}]")
        p(f"  {'─'*80}")
        p(f"    SUBTOTAL AVOIDABLE: {avoidable_total:,.0f} NIS")

    # Show subscriptions
    sub_items = [t for t in classified if t['cls'] == 'subscription']
    if sub_items:
        sub_items.sort(key=lambda x: -x['amount'])
        p(f"\n  SUBSCRIPTIONS (review if needed):")
        p(f"  {'─'*80}")
        for t in sub_items:
            adv = f" -> {t['advice']}" if t['advice'] else ""
            p(f"    {t['date'].strftime('%d/%m')}  {t['amount']:>8,.0f} NIS  {t['name'][:45]:<45s}{adv}")
        p(f"  {'─'*80}")
        p(f"    SUBTOTAL SUBSCRIPTIONS: {subscription_total:,.0f} NIS")

    # Savings estimate
    potential_savings = avoidable_total * 0.6 + subscription_total * 0.3
    p(f"\n  POTENTIAL MONTHLY SAVINGS: ~{potential_savings:,.0f} NIS")
    p(f"    (Cutting 60% of avoidable + 30% of subscriptions)")

p(f"\n{'='*90}")
p(f"  GRAND SUMMARY (Oct 2025 - Feb 2026)")
p(f"{'='*90}")
p(f"  Total spent:                    {grand_total:,.0f} NIS")
p(f"  Total avoidable:                {grand_avoidable:,.0f} NIS")
p(f"  Total subscriptions:            {grand_subscription:,.0f} NIS")
p(f"  Combined avoidable+subs:        {grand_avoidable + grand_subscription:,.0f} NIS")
p(f"")
p(f"  REALISTIC SAVINGS POTENTIAL:")
p(f"    If you cut 60% of avoidable:  ~{grand_avoidable * 0.6:,.0f} NIS")
p(f"    If you trim subscriptions:    ~{grand_subscription * 0.3:,.0f} NIS")
p(f"    TOTAL POTENTIAL SAVINGS:      ~{grand_avoidable * 0.6 + grand_subscription * 0.3:,.0f} NIS over 5 months")
avg_monthly = (grand_avoidable * 0.6 + grand_subscription * 0.3) / 5
p(f"    AVG MONTHLY SAVINGS:          ~{avg_monthly:,.0f} NIS/month")

p(f"\n{'='*90}")
p(f"  TOP SAVINGS RECOMMENDATIONS")
p(f"{'='*90}")

# Find recurring avoidable patterns
all_classified = []
for m in focus_months:
    if m not in by_month:
        continue
    for t in by_month[m]:
        cls, sub_cls, advice = classify(t)
        all_classified.append({**t, 'cls': cls, 'sub_cls': sub_cls})

# Group avoidable by sub_cls
avoidable_by_cat = defaultdict(lambda: {'total': 0, 'count': 0, 'items': []})
for t in all_classified:
    if t['cls'] in ('avoidable', 'subscription'):
        avoidable_by_cat[t['sub_cls']]['total'] += t['amount']
        avoidable_by_cat[t['sub_cls']]['count'] += 1
        avoidable_by_cat[t['sub_cls']]['items'].append(t)

for cat, data in sorted(avoidable_by_cat.items(), key=lambda x: -x[1]['total']):
    p(f"\n  {cat.upper()} - Total: {data['total']:,.0f} NIS ({data['count']} transactions)")
    for item in sorted(data['items'], key=lambda x: -x['amount'])[:5]:
        p(f"    {item['date'].strftime('%d/%m/%Y')}  {item['amount']:>8,.0f} NIS  {item['name'][:55]}")

# Write to file
with open('spending_analysis.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))

# Also print
for line in output:
    print(line)
